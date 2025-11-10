import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io

# Page configuration
st.set_page_config(
    page_title="Feasibility Analysis Tool",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Enhanced CSS for better styling and performance
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        padding: 1rem 0;
    }
    .section-header {
        font-size: 1.5rem;
        font-weight: bold;
        color: #2c3e50;
        margin-top: 1rem;
        margin-bottom: 0.5rem;
        border-bottom: 2px solid #1f77b4;
        padding-bottom: 0.3rem;
    }
    .metric-card {
        background-color: #f0f2f6;
        padding: 1rem;
        border-radius: 0.5rem;
        margin: 0.5rem 0;
    }
    .stButton>button {
        width: 100%;
    }
    .data-editor {
        background-color: white;
        border: 1px solid #ddd;
        border-radius: 0.5rem;
        padding: 0.5rem;
    }
    .section-title {
        font-weight: bold;
        font-size: 1.1rem;
        color: #34495e;
        margin: 0.5rem 0;
    }
    .item-row {
        border-bottom: 1px solid #ecf0f1;
        padding: 0.5rem 0;
    }
    .delete-btn {
        background-color: #e74c3c !important;
        border: none !important;
        color: white !important;
    }
    .add-btn {
        background-color: #27ae60 !important;
        border: none !important;
        color: white !important;
    }
    /* Hide streamlit's auto-save button */
    .stFormSubmitButton button {
        display: none;
    }
    </style>
    """, unsafe_allow_html=True)

# Initialize session state with enhanced structure
def init_session_state():
    # Flag to track if we've already initialized in this session
    if 'initialized' not in st.session_state:
        st.session_state.initialized = False

    # Only initialize once per session
    if not st.session_state.initialized:
        # Try to load from persistent storage FIRST
        loaded = False
        try:
            # Import here to avoid circular dependency
            import os
            STORAGE_FILE_CHECK = "feasibilitizer_data.json"
            if os.path.exists(STORAGE_FILE_CHECK):
                with open(STORAGE_FILE_CHECK, 'r') as f:
                    data = json.load(f)

                st.session_state.capex_items = data.get('capex_items', [])
                st.session_state.opex_cash_in = data.get('opex_cash_in', [])
                st.session_state.opex_cash_out = data.get('opex_cash_out', [])
                st.session_state.project_years = int(data.get('project_years', 5))  # Ensure integer
                st.session_state.discount_rate = float(data.get('discount_rate', 12.0))
                st.session_state.default_data_loaded = data.get('default_data_loaded', True)
                st.session_state.last_save = datetime.now()
                loaded = True
        except:
            loaded = False

        # If no saved data found, initialize with defaults
        if not loaded:
            st.session_state.capex_items = []
            st.session_state.opex_cash_in = []
            st.session_state.opex_cash_out = []
            st.session_state.project_years = 5
            st.session_state.discount_rate = 12.0
            st.session_state.last_save = datetime.now()
            st.session_state.default_data_loaded = False

        # Growth rates - UI setting only, NOT saved with data
        if 'opex_in_growth' not in st.session_state:
            st.session_state.opex_in_growth = 0.0
        if 'opex_out_growth' not in st.session_state:
            st.session_state.opex_out_growth = 0.0

            # Add default data only on first load
            st.session_state.capex_items = [
                {"id": "79e1c473-4e21-4ac3-af5f-99b6e0cbfc73", "name": "Synology NAS Server", "volume": 1.0, "unit": "unit", "price": 10599000.0},
                {"id": "bf7935f4-6824-471a-a7d0-d1c0b72ea194", "name": "Uninterruptible Power Supply (UPS)", "volume": 1.0, "unit": "unit", "price": 3529000.0},
                {"id": "51d86d79-43d9-41a9-a6ec-a106c839f2ef", "name": "Network Switch", "volume": 2.0, "unit": "unit", "price": 132900.0},
                {"id": "4157ff5c-8e94-4e15-9431-974da24ef6da", "name": "Laptop/Desktop", "volume": 9.0, "unit": "unit", "price": 6671000.0},
                {"id": "11616bc0-3e98-4618-a9f0-aaefaec93fe4", "name": "Tablet", "volume": 6.0, "unit": "unit", "price": 4249150.0},
                {"id": "303585da-d189-4973-a946-f8847035de87", "name": "Wireless Access Point", "volume": 3.0, "unit": "unit", "price": 175000.0},
                {"id": "240d5d78-811b-4717-9034-4ea76e27e32d", "name": "Biaya Pengembangan Sistem", "volume": 1.0, "unit": "paket", "price": 28000000.0},
                {"id": "af543c24-d7ad-423b-8bed-cac90c8b8270", "name": "Biaya Setup & Instalasi", "volume": 1.0, "unit": "paket", "price": 2000000.0},
                {"id": "5fb9ae44-e5bc-47d9-869f-d90d12a69e41", "name": "Biaya Onboarding & Training", "volume": 1.0, "unit": "paket", "price": 1500000.0},
                {"id": "b5f2f700-b75b-4db9-81f1-802fe06f6051", "name": "Biaya Domain & Konfigurasi", "volume": 1.0, "unit": "paket", "price": 2319900.0}
            ]

            st.session_state.opex_cash_in = [
                {"id": "cf693797-5cd7-4985-9245-a3f96e562162", "name": "Penghematan biaya tenaga kerja administrasi", "volume": 1.0, "unit": "paket", "price": 21600000.0},
                {"id": "631868b8-1aad-46dd-93ed-b7522534f514", "name": "Pengurangan biaya kesalahan & rework", "volume": 1.0, "unit": "paket", "price": 16800000.0},
                {"id": "471147a2-d322-40c7-a45d-293b806b6fde", "name": "Peningkatan produktivitas proses bisnis", "volume": 1.0, "unit": "paket", "price": 45153948.0},
                {"id": "eeaa72cc-023e-438a-bdc6-4b8f724f91ee", "name": "Penghematan biaya dokumen fisik", "volume": 1.0, "unit": "paket", "price": 4813680.0}
            ]

            st.session_state.opex_cash_out = [
                {"id": "278e5ae1-b76f-4601-9f19-1cc84079b9d4", "name": "Koneksi Internet Dedicated", "volume": 2.0, "unit": "paket", "price": 375000.0},
                {"id": "75090947-0595-4e0d-ad68-3df78a71b876", "name": "Listrik (Server & Infrastruktur) PLN tarif R1/900VA", "volume": 1.0, "unit": "paket", "price": 456000.0},
                {"id": "de1e3463-a3cb-46f1-9b49-f73ea19a9f21", "name": "IP Public Cloudflare", "volume": 1.0, "unit": "paket", "price": 331440.0},
                {"id": "8ebb7c15-4991-4e94-9db0-59473729d032", "name": "Maintenance & Support Teknis", "volume": 1.0, "unit": "paket", "price": 1500000.0}
            ]

            st.session_state.default_data_loaded = True

        # Mark as initialized
        st.session_state.initialized = True

init_session_state()

# Persistent storage functions using local file
STORAGE_FILE = "feasibilitizer_data.json"

def save_to_storage():
    """Save current state to persistent storage file"""
    try:
        data = {
            'capex_items': st.session_state.capex_items,
            'opex_cash_in': st.session_state.opex_cash_in,
            'opex_cash_out': st.session_state.opex_cash_out,
            'project_years': int(st.session_state.project_years),  # Ensure integer
            'discount_rate': float(st.session_state.discount_rate),
            'last_save': st.session_state.last_save.isoformat(),
            'default_data_loaded': st.session_state.default_data_loaded
        }

        # Save to file
        with open(STORAGE_FILE, 'w') as f:
            json.dump(data, f, indent=2)

        return data
    except Exception as e:
        # Silently fail - don't show error to user for auto-save
        return None

def load_from_storage():
    """Load data from persistent storage file"""
    try:
        import os
        if os.path.exists(STORAGE_FILE):
            with open(STORAGE_FILE, 'r') as f:
                data = json.load(f)

            st.session_state.capex_items = data.get('capex_items', [])
            st.session_state.opex_cash_in = data.get('opex_cash_in', [])
            st.session_state.opex_cash_out = data.get('opex_cash_out', [])
            st.session_state.project_years = int(data.get('project_years', 5))  # Ensure integer
            st.session_state.discount_rate = float(data.get('discount_rate', 12.0))
            st.session_state.default_data_loaded = data.get('default_data_loaded', False)

            if 'last_save' in data:
                st.session_state.last_save = datetime.fromisoformat(data['last_save'])

            return True
        return False
    except Exception as e:
        # If file is corrupted or unreadable, just return False
        return False

# Auto-save function
def auto_save():
    st.session_state.last_save = datetime.now()
    save_to_storage()

# Helper Functions with improved error handling
def calculate_total(item):
    """Calculate total from volume * price with validation"""
    try:
        volume = float(item.get('volume', 0))
        price = float(item.get('price', 0))
        return volume * price
    except (ValueError, TypeError):
        return 0

def calculate_capex_total():
    """Calculate total CAPEX with validation"""
    return sum(calculate_total(item) for item in st.session_state.capex_items)

def calculate_yearly_opex_cash_in():
    """Calculate yearly OPEX cash in with validation"""
    return sum(calculate_total(item) for item in st.session_state.opex_cash_in)

def calculate_yearly_opex_cash_out():
    """Calculate yearly OPEX cash out with validation"""
    return sum(calculate_total(item) for item in st.session_state.opex_cash_out)

def calculate_net_cashflow():
    """Calculate net cashflow for each year with growth rates applied for display only"""
    try:
        years = int(st.session_state.project_years)  # Ensure integer
        capex = calculate_capex_total()

        # Base values from input data (never changed!)
        base_cash_in = calculate_yearly_opex_cash_in()
        base_cash_out = calculate_yearly_opex_cash_out()

        # Get growth rates (convert % to decimal)
        growth_in = st.session_state.opex_in_growth / 100.0
        growth_out = st.session_state.opex_out_growth / 100.0

        cashflows = [-capex]  # Year 0

        # Calculate each year with compound growth
        for year in range(1, years + 1):
            # Apply growth: Amount = Base × (1 + rate)^year
            cash_in_year = base_cash_in * ((1 + growth_in) ** year)
            cash_out_year = base_cash_out * ((1 + growth_out) ** year)
            yearly_net = cash_in_year - cash_out_year
            cashflows.append(yearly_net)

        return cashflows
    except Exception:
        return [0]

def calculate_discount_factor():
    """Calculate discount factors for each year"""
    try:
        years = int(st.session_state.project_years)  # Ensure integer
        rate = st.session_state.discount_rate / 100
        factors = [1.0]  # Year 0
        for i in range(1, years + 1):
            factors.append((1 + rate) ** i)
        return factors
    except Exception:
        return [1.0]

def calculate_discounted_cashflow(cashflows, discount_factors):
    """Calculate discounted cashflow"""
    try:
        return [cf / df if df != 0 else 0 for cf, df in zip(cashflows, discount_factors)]
    except Exception:
        return [0]

def calculate_cumulative_cashflow(discounted_cashflows):
    """Calculate cumulative cashflow"""
    try:
        cumulative = []
        total = 0
        for dcf in discounted_cashflows:
            total += dcf
            cumulative.append(total)
        return cumulative
    except Exception:
        return [0]

def calculate_npv(discounted_cashflows):
    """Calculate Net Present Value"""
    try:
        return sum(discounted_cashflows)
    except Exception:
        return 0

def calculate_irr(cashflows):
    """Calculate Internal Rate of Return"""
    try:
        from numpy_financial import irr
        return irr(cashflows)
    except:
        # Fallback to manual calculation if numpy_financial not available
        try:
            return np.irr(cashflows)
        except:
            return 0.0

def calculate_payback_period(cumulative_cashflows):
    """Calculate Payback Period"""
    try:
        for i, cumulative in enumerate(cumulative_cashflows):
            if cumulative >= 0:
                if i == 0:
                    return 0
                # Linear interpolation for more accurate payback period
                prev_cumulative = cumulative_cashflows[i-1]
                curr_cumulative = cumulative
                if curr_cumulative - prev_cumulative != 0:
                    fraction = abs(prev_cumulative) / (curr_cumulative - prev_cumulative)
                    return i - 1 + fraction
        return st.session_state.project_years  # If never positive
    except Exception:
        return 0

# Excel export function with formatting
def export_to_excel():
    """Export cash flow analysis to Excel with formatting"""
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Cash Flow Analysis"

        # Calculate all values
        cashflows = calculate_net_cashflow()
        discount_factors = calculate_discount_factor()
        discounted_cashflows = calculate_discounted_cashflow(cashflows, discount_factors)
        cumulative_cashflows = calculate_cumulative_cashflow(discounted_cashflows)

        # Create headers
        years_labels = ['Tahun 0'] + [f'Tahun {i+1}' for i in range(int(st.session_state.project_years))]
        headers = ['Description', 'Total'] + years_labels

        # Add title
        ws.merge_cells('A1:' + chr(65 + len(headers) - 1) + '1')
        title_cell = ws['A1']
        title_cell.value = "CASH FLOW ANALYSIS (ARUS KAS)"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        title_cell.fill = PatternFill(start_color="1f77b4", end_color="1f77b4", fill_type="solid")
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")

        # Add project info
        ws.merge_cells('A2:' + chr(65 + len(headers) - 1) + '2')
        ws['A2'].value = f"Project Duration: {st.session_state.project_years} years | Discount Rate: {st.session_state.discount_rate}%"
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(size=10, color="666666")

        # Add timestamp
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.merge_cells('A3:' + chr(65 + len(headers) - 1) + '3')
        ws['A3'].value = f"Generated: {timestamp}"
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A3'].font = Font(size=10, color="999999")

        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="ecf0f1", end_color="ecf0f1", fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

        current_row = 6

        # Function to add section with color
        def add_section(title, data, color="FFFFFF"):
            nonlocal current_row

            # Add section title
            ws.merge_cells(f'A{current_row}:{chr(65 + len(headers) - 1)}{current_row}')
            title_cell = ws.cell(row=current_row, column=1, value=title)
            title_cell.font = Font(bold=True, size=12)
            title_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center')
            current_row += 1

            # Add data rows
            for row_data in data:
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=value)
                    cell.border = Border(
                        left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin')
                    )
                    if isinstance(value, (int, float)) and value != '':
                        cell.number_format = '#,##0'
                current_row += 1

            current_row += 1  # Add spacing

        # CAPEX Section (Light Blue)
        capex_data = []
        capex_total = 0
        for item in st.session_state.capex_items:
            total = -calculate_total(item)
            capex_total += total
            row = [f"  {item['name']}", total]
            for i, year in enumerate(years_labels):
                if i == 0:
                    row.append(total)
                else:
                    row.append('')
            capex_data.append(row)

        capex_data.append(['Total CAPEX', capex_total] + ([capex_total] + [''] * (len(years_labels) - 1)))
        add_section('CAPITAL EXPENDITURE (CAPEX)', capex_data, "DDEBF7")

        # OPEX Cash In Section (Light Green)
        cashin_data = []
        growth_in = st.session_state.opex_in_growth / 100.0
        for item in st.session_state.opex_cash_in:
            base_total = calculate_total(item)
            row = [f"  {item['name']}", base_total]
            for i, year in enumerate(years_labels):
                if i == 0:
                    row.append('')
                else:
                    # Apply growth: base × (1 + rate)^year
                    year_total = base_total * ((1 + growth_in) ** i)
                    row.append(year_total)
            cashin_data.append(row)

        base_yearly_revenue = calculate_yearly_opex_cash_in()
        revenue_row = ['Total Revenue', '']
        for i, year in enumerate(years_labels):
            if i == 0:
                revenue_row.append('')
            else:
                year_revenue = base_yearly_revenue * ((1 + growth_in) ** i)
                revenue_row.append(year_revenue)
        cashin_data.append(revenue_row)
        add_section('OPERATIONAL REVENUE (OPEX - Cash In)', cashin_data, "D4E6C4")

        # OPEX Cash Out Section (Light Red)
        cashout_data = []
        growth_out = st.session_state.opex_out_growth / 100.0
        for item in st.session_state.opex_cash_out:
            base_total = -calculate_total(item)
            row = [f"  {item['name']}", base_total]
            for i, year in enumerate(years_labels):
                if i == 0:
                    row.append('')
                else:
                    # Apply growth: base × (1 + rate)^year
                    year_total = base_total * ((1 + growth_out) ** i)
                    row.append(year_total)
            cashout_data.append(row)

        base_yearly_expenses = -calculate_yearly_opex_cash_out()
        expenses_row = ['Total Expenses', '']
        for i, year in enumerate(years_labels):
            if i == 0:
                expenses_row.append('')
            else:
                year_expenses = base_yearly_expenses * ((1 + growth_out) ** i)
                expenses_row.append(year_expenses)
        cashout_data.append(expenses_row)
        add_section('OPERATIONAL EXPENSES (OPEX - Cash Out)', cashout_data, "F8D7DA")

        # Financial Summary Section (Light Yellow)
        summary_data = [
            ['NET CASH FLOW', ''] + cashflows,
            [f'DISCOUNT FACTOR (MARR {st.session_state.discount_rate}%)', ''] + discount_factors,
            ['DISCOUNTED CASH FLOW', ''] + discounted_cashflows,
            ['CUMULATIVE CASH FLOW', ''] + cumulative_cashflows,
            ['NPV', calculate_npv(discounted_cashflows)] + [''] * len(years_labels),
            ['IRR', f"{calculate_irr(cashflows)*100:.2f}%"] + [''] * len(years_labels),
            ['Payback Period', f"{calculate_payback_period(cumulative_cashflows):.2f} years"] + [''] * len(years_labels)
        ]
        add_section('FINANCIAL SUMMARY', summary_data, "FFF2CC")

        # Adjust column widths
        for col in range(1, len(headers) + 1):
            if col == 1:  # Description column
                ws.column_dimensions[chr(64 + col)].width = 40
            else:
                ws.column_dimensions[chr(64 + col)].width = 15

        # Save to bytes
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        return excel_file
    except Exception as e:
        st.error(f"Error exporting to Excel: {str(e)}")
        return None

# Enhanced input component with auto-save
def editable_data_editor(section_title, items_key, color="primary", unit_placeholder="unit"):
    """Enhanced data editor with auto-save and better UX"""

    # Get reference to items
    items = st.session_state[items_key]

    # Create proper unique key for this section
    item_key = items_key  # Use the full key to avoid conflicts

    # Use the provided section title
    st.markdown(f"### {section_title}")

    # Bulk paste feature
    with st.expander("📋 Bulk Add from Clipboard", expanded=False):
        st.markdown("**Paste data from Excel/Google Sheets** (columns: Name | Qty | Unit | Price)")
        bulk_paste = st.text_area("Paste rows here (tab-separated)", key=f"bulk_{item_key}", height=150,
                                   help="Copy rows from Excel/Sheets and paste here. Format: Name [Tab] Qty [Tab] Unit [Tab] Price")

        if st.button("📥 Import All Rows", key=f"bulk_add_{item_key}", type="primary"):
            if bulk_paste.strip():
                try:
                    lines = bulk_paste.strip().split('\n')
                    added_count = 0
                    import uuid

                    for line in lines:
                        # Skip empty lines
                        if not line.strip():
                            continue

                        # Try tab-separated first, then comma-separated
                        if '\t' in line:
                            parts = line.split('\t')
                        else:
                            parts = [p.strip() for p in line.split(',')]

                        if len(parts) >= 4:
                            try:
                                name = parts[0].strip()
                                volume = float(parts[1].strip())
                                unit = parts[2].strip()
                                # Remove any currency symbols and commas from price
                                price_str = parts[3].strip().replace('Rp', '').replace(',', '').replace('.', '')
                                price = float(price_str)

                                if name and volume > 0 and price > 0:
                                    st.session_state[items_key].append({
                                        "id": str(uuid.uuid4()),
                                        "name": name,
                                        "volume": volume,
                                        "unit": unit,
                                        "price": price
                                    })
                                    added_count += 1
                            except (ValueError, IndexError):
                                continue  # Skip invalid rows

                    if added_count > 0:
                        auto_save()
                        st.success(f"✅ Added {added_count} items successfully!")
                        st.rerun()
                    else:
                        st.error("❌ No valid rows found. Please check format: Name | Qty | Unit | Price")
                except Exception as e:
                    st.error(f"❌ Error importing: {str(e)}")
            else:
                st.warning("⚠️ Please paste some data first")

    # Add new item form
    with st.expander("➕ Add New Item", expanded=False):
        col1, col2, col3, col4, col5 = st.columns([3, 1, 2, 2, 1])

        with col1:
            new_name = st.text_input("Item Name", key=f"new_{item_key}_name")
        with col2:
            new_volume = st.number_input("Qty", min_value=0.0, value=1.0, key=f"new_{item_key}_volume")
        with col3:
            new_unit = st.text_input("Unit", value=unit_placeholder, key=f"new_{item_key}_unit")
        with col4:
            new_price = st.number_input("Price", min_value=0.0, value=0.0, step=10000.0, key=f"new_{item_key}_price")
        with col5:
            if st.button("Add", key=f"add_{item_key}", type="primary"):
                if new_name and new_volume > 0 and new_price > 0:
                    import uuid
                    st.session_state[items_key].append({
                        "id": str(uuid.uuid4()),
                        "name": new_name,
                        "volume": new_volume,
                        "unit": new_unit,
                        "price": new_price
                    })
                    auto_save()
                    st.success(f"✅ Added: {new_name}")
                    st.rerun()
                else:
                    st.error("❌ Please fill all required fields")

    # Display and edit existing items
    if st.session_state[items_key]:  # Check session state directly
        # Show title first (total will be calculated after widgets render)
        st.markdown(f'<div class="section-title">Current Items</div>', unsafe_allow_html=True)

        # Track live total as we render items
        live_total = 0.0

        # Loop through items - enumerate to get both index and item
        for idx, item in enumerate(st.session_state[items_key]):
            # Ensure item has an ID (for backward compatibility with old data)
            if 'id' not in item:
                import uuid
                item['id'] = str(uuid.uuid4())

            item_id = item['id']  # Use item ID for widget keys!

            with st.container():
                col1, col2, col3, col4, col5, col6 = st.columns([3, 1, 1.5, 2, 2, 1.5])

                with col1:
                    name = st.text_input("Name", value=item.get('name', ''), key=f"{item_key}_name_{item_id}", label_visibility="collapsed")

                with col2:
                    volume = st.number_input("Qty", value=float(item.get('volume', 0)), min_value=0.0, key=f"{item_key}_vol_{item_id}", label_visibility="collapsed")

                with col3:
                    unit = st.text_input("Unit", value=item.get('unit', ''), key=f"{item_key}_unit_{item_id}", label_visibility="collapsed")

                with col4:
                    price = st.number_input("Price", value=float(item.get('price', 0)), min_value=0.0, step=10000.0, key=f"{item_key}_price_{item_id}", label_visibility="collapsed")

                with col5:
                    # Recalculate total from current widget values
                    current_total = float(volume) * float(price)
                    live_total += current_total  # Add to running total
                    st.markdown(f'<div style="text-align: right; font-weight: bold; padding-top: 8px;">Rp {current_total:,.0f}</div>', unsafe_allow_html=True)

                with col6:
                    btn_col1, btn_col2 = st.columns(2)
                    with btn_col1:
                        if st.button("📋", key=f"dup_{item_key}_{item_id}", help="Duplicate item"):
                            import uuid
                            # Create a copy with new ID
                            duplicated_item = {
                                "id": str(uuid.uuid4()),
                                "name": name,
                                "volume": volume,
                                "unit": unit,
                                "price": price
                            }
                            st.session_state[items_key].insert(idx + 1, duplicated_item)
                            auto_save()
                            st.rerun()
                    with btn_col2:
                        if st.button("🗑️", key=f"del_{item_key}_{item_id}", help="Delete item"):
                            # Delete by finding the item with this ID
                            st.session_state[items_key] = [i for i in st.session_state[items_key] if i.get('id') != item_id]
                            auto_save()
                            st.rerun()

            # Update session state ONLY if values changed (prevents unintended overwrites)
            changed = False
            if st.session_state[items_key][idx]['name'] != name:
                st.session_state[items_key][idx]['name'] = name
                changed = True
            if st.session_state[items_key][idx]['volume'] != volume:
                st.session_state[items_key][idx]['volume'] = volume
                changed = True
            if st.session_state[items_key][idx]['unit'] != unit:
                st.session_state[items_key][idx]['unit'] = unit
                changed = True
            if st.session_state[items_key][idx]['price'] != price:
                st.session_state[items_key][idx]['price'] = price
                changed = True

            # Only auto-save if something actually changed
            if changed:
                auto_save()

        # Display the live total after all items rendered
        st.markdown("---")
        st.markdown(f'<div style="text-align: right; font-size: 1.2em; font-weight: bold; padding: 10px; background-color: #f0f2f6; border-radius: 5px;">TOTAL: Rp {live_total:,.0f}</div>', unsafe_allow_html=True)

    else:
        st.info("No items added yet. Click 'Add New Item' to start.")

# Main App
st.markdown('<div class="main-header">📊 Feasibility Analysis Tool</div>', unsafe_allow_html=True)

# Auto-save indicator
last_save_str = st.session_state.last_save.strftime("%H:%M:%S")
st.markdown(f'<div style="text-align: right; color: #666; font-size: 0.8rem;">🔄 Auto-saved at {last_save_str}</div>', unsafe_allow_html=True)

st.markdown("---")

# Enhanced Sidebar Configuration
with st.sidebar:
    st.header("⚙️ Project Configuration")

    st.subheader("General Settings")

    # Fixed project duration input with better UX
    new_years = st.number_input(
        "Project Duration (Years)",
        min_value=1.0,
        max_value=30.0,
        value=float(st.session_state.project_years),
        step=1.0,
        key="project_years_input"
    )

    if new_years != st.session_state.project_years:
        st.session_state.project_years = new_years
        auto_save()
        st.rerun()

    # Fixed discount rate input with better UX
    new_rate = st.number_input(
        "Discount Rate (MARR %)",
        min_value=0.0,
        max_value=100.0,
        value=float(st.session_state.discount_rate),
        step=0.1,
        key="discount_rate_input"
    )

    if abs(new_rate - st.session_state.discount_rate) > 0.01:
        st.session_state.discount_rate = new_rate
        auto_save()
        st.rerun()

    st.markdown("---")

    # Growth rates for analysis display only
    st.subheader("📈 Yearly Growth Rates")
    st.markdown("**For analysis display only** (doesn't change input data)")
    st.info("💡 These growth rates use compound interest formula: Amount = Base × (1 + rate)^year")

    col1, col2 = st.columns(2)

    with col1:
        st.session_state.opex_in_growth = st.number_input(
            "💰 Cash In / Revenue Growth (%/year)",
            min_value=-100.0,
            max_value=100.0,
            value=float(st.session_state.opex_in_growth),
            step=0.5,
            key="opex_in_growth_input",
            help="Annual growth rate applied in Cash Flow Analysis tab (compound growth)"
        )

    with col2:
        st.session_state.opex_out_growth = st.number_input(
            "💸 Cash Out / Expense Growth (%/year)",
            min_value=-100.0,
            max_value=100.0,
            value=float(st.session_state.opex_out_growth),
            step=0.5,
            key="opex_out_growth_input",
            help="Annual growth rate applied in Cash Flow Analysis tab (compound growth)"
        )

    # Show example calculation
    with st.expander("📊 See Growth Calculation Example"):
        base_revenue = calculate_yearly_opex_cash_in()
        if base_revenue > 0:
            years_demo = 3
            growth_demo = st.session_state.opex_in_growth / 100.0
            st.write("**Revenue Growth Example:**")
            for year in range(years_demo + 1):
                revenue = base_revenue * ((1 + growth_demo) ** year)
                st.write(f"Year {year}: Rp {revenue:,.0f}")
        else:
            st.write("Add revenue items to see growth calculation example")

    st.markdown("---")

    # Export/Import Data
    st.subheader("💾 Data Management")

    # Save Project
    st.markdown("**Save Your Work:**")
    project_data = save_to_storage()
    if project_data:
        st.download_button(
            label="💾 Save Project as JSON",
            data=json.dumps(project_data, indent=2),
            file_name=f"project_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json",
            mime="application/json",
            help="Download project data to load later",
            use_container_width=True
        )

    st.markdown("---")

    # Load Project
    st.markdown("**Load Previous Work:**")
    uploaded_file = st.file_uploader("Choose JSON file", type=['json'], label_visibility="collapsed", key="project_upload")
    if uploaded_file is not None:
        try:
            loaded_data = json.loads(uploaded_file.read())
            st.session_state.capex_items = loaded_data.get('capex_items', [])
            st.session_state.opex_cash_in = loaded_data.get('opex_cash_in', [])
            st.session_state.opex_cash_out = loaded_data.get('opex_cash_out', [])
            st.session_state.project_years = int(loaded_data.get('project_years', 5))
            st.session_state.discount_rate = float(loaded_data.get('discount_rate', 12.0))
            auto_save()
            st.success("✅ Project loaded successfully!")
            st.rerun()
        except Exception as e:
            st.error(f"❌ Error loading project: {str(e)}")

    st.markdown("---")

    # Export to Excel
    st.markdown("**Export Analysis:**")
    if st.button("📥 Export to Excel", type="primary", use_container_width=True):
        excel_file = export_to_excel()
        if excel_file:
            st.download_button(
                label="⬇️ Download Excel File",
                data=excel_file,
                file_name=f"analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    st.markdown("---")

    # Reset
    if st.button("🔄 Reset to Default", help="Reset all data to example values", use_container_width=True):
        # Reset all data
        st.session_state.capex_items = []
        st.session_state.opex_cash_in = []
        st.session_state.opex_cash_out = []
        init_session_state()
        auto_save()
        st.success("✅ Reset to default values")
        st.rerun()

    # Add project summary
    st.markdown("---")
    st.subheader("📊 Quick Summary")

    capex_total = calculate_capex_total()
    yearly_revenue = calculate_yearly_opex_cash_in()
    yearly_expenses = calculate_yearly_opex_cash_out()

    st.metric("CAPEX", f"Rp {capex_total:,.0f}")
    st.metric("Annual Revenue", f"Rp {yearly_revenue:,.0f}")
    st.metric("Annual Expenses", f"Rp {yearly_expenses:,.0f}")
    st.metric("Net Cash Flow/Year", f"Rp {yearly_revenue - yearly_expenses:,.0f}")

# Main Tabs
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📝 Input Data",
    "📊 Cash Flow Analysis",
    "💰 Financial Metrics",
    "📈 Visualizations",
    "🎯 Sensitivity Analysis"
])

# TAB 1: INPUT DATA (Enhanced)
with tab1:
    st.markdown('<div class="section-header">Input Data Management</div>', unsafe_allow_html=True)

    # Enhanced input sections
    editable_data_editor("💼 Capital Expenditure (CAPEX)", "capex_items", "primary", "unit")
    st.markdown("---")
    editable_data_editor("💵 Operating Revenue (OPEX - Cash In)", "opex_cash_in", "success", "month")
    st.markdown("---")
    editable_data_editor("💸 Operating Expenses (OPEX - Cash Out)", "opex_cash_out", "danger", "month")

# TAB 2: CASH FLOW ANALYSIS (Enhanced)
with tab2:
    st.markdown('<div class="section-header">Cash Flow Analysis (Arus Kas)</div>', unsafe_allow_html=True)

    # Calculate all values
    cashflows = calculate_net_cashflow()
    discount_factors = calculate_discount_factor()
    discounted_cashflows = calculate_discounted_cashflow(cashflows, discount_factors)
    cumulative_cashflows = calculate_cumulative_cashflow(discounted_cashflows)

    # Create comprehensive cash flow table
    years_labels = ['Tahun 0'] + [f'Tahun {i+1}' for i in range(int(st.session_state.project_years))]

    # Build detailed breakdown
    cash_flow_data = {
        'Description': [],
        'Total': []
    }

    for year in years_labels:
        cash_flow_data[year] = []

    # CAPEX Section
    cash_flow_data['Description'].append('CAPITAL EXPENDITURE (CAPEX)')
    cash_flow_data['Total'].append('')
    for i, year in enumerate(years_labels):
        cash_flow_data[year].append('')

    capex_total = 0
    for item in st.session_state.capex_items:
        total = -calculate_total(item)  # Negative for expenses
        capex_total += total
        cash_flow_data['Description'].append(f"  {item['name']}")
        cash_flow_data['Total'].append(total)
        for i, year in enumerate(years_labels):
            if i == 0:
                cash_flow_data[year].append(total)
            else:
                cash_flow_data[year].append('')

    cash_flow_data['Description'].append('Total CAPEX')
    cash_flow_data['Total'].append(capex_total)
    for i, year in enumerate(years_labels):
        if i == 0:
            cash_flow_data[year].append(capex_total)
        else:
            cash_flow_data[year].append('')

    # Empty row
    cash_flow_data['Description'].append('')
    cash_flow_data['Total'].append('')
    for year in years_labels:
        cash_flow_data[year].append('')

    # OPEX Cash In Section
    cash_flow_data['Description'].append('OPERATIONAL REVENUE (OPEX - Cash In)')
    cash_flow_data['Total'].append('')
    for year in years_labels:
        cash_flow_data[year].append('')

    # Get growth rate
    growth_in = st.session_state.opex_in_growth / 100.0

    for item in st.session_state.opex_cash_in:
        base_total = calculate_total(item)
        cash_flow_data['Description'].append(f"  {item['name']}")
        cash_flow_data['Total'].append(base_total)
        for i, year in enumerate(years_labels):
            if i == 0:
                cash_flow_data[year].append('')
            else:
                # Apply growth: base × (1 + rate)^year
                year_total = base_total * ((1 + growth_in) ** i)
                cash_flow_data[year].append(year_total)

    cash_flow_data['Description'].append('Total Revenue')
    base_yearly_revenue = calculate_yearly_opex_cash_in()
    cash_flow_data['Total'].append('')
    for i, year in enumerate(years_labels):
        if i == 0:
            cash_flow_data[year].append('')
        else:
            # Apply growth: base × (1 + rate)^year
            year_revenue = base_yearly_revenue * ((1 + growth_in) ** i)
            cash_flow_data[year].append(year_revenue)

    # Empty row
    cash_flow_data['Description'].append('')
    cash_flow_data['Total'].append('')
    for year in years_labels:
        cash_flow_data[year].append('')

    # OPEX Cash Out Section
    cash_flow_data['Description'].append('OPERATIONAL EXPENSES (OPEX - Cash Out)')
    cash_flow_data['Total'].append('')
    for year in years_labels:
        cash_flow_data[year].append('')

    # Get growth rate for expenses
    growth_out = st.session_state.opex_out_growth / 100.0

    for item in st.session_state.opex_cash_out:
        base_total = -calculate_total(item)  # Negative for expenses
        cash_flow_data['Description'].append(f"  {item['name']}")
        cash_flow_data['Total'].append(base_total)
        for i, year in enumerate(years_labels):
            if i == 0:
                cash_flow_data[year].append('')
            else:
                # Apply growth: base × (1 + rate)^year
                year_total = base_total * ((1 + growth_out) ** i)
                cash_flow_data[year].append(year_total)

    cash_flow_data['Description'].append('Total Expenses')
    base_yearly_expenses = -calculate_yearly_opex_cash_out()
    cash_flow_data['Total'].append('')
    for i, year in enumerate(years_labels):
        if i == 0:
            cash_flow_data[year].append('')
        else:
            # Apply growth: base × (1 + rate)^year
            year_expenses = base_yearly_expenses * ((1 + growth_out) ** i)
            cash_flow_data[year].append(year_expenses)

    # Empty row
    cash_flow_data['Description'].append('')
    cash_flow_data['Total'].append('')
    for year in years_labels:
        cash_flow_data[year].append('')

    # Net Cash Flow
    cash_flow_data['Description'].append('NET CASH FLOW')
    cash_flow_data['Total'].append('')
    for i, year in enumerate(years_labels):
        cash_flow_data[year].append(cashflows[i])

    # Empty row
    cash_flow_data['Description'].append('')
    cash_flow_data['Total'].append('')
    for year in years_labels:
        cash_flow_data[year].append('')

    # Discount Factor
    rate_percent = st.session_state.discount_rate
    cash_flow_data['Description'].append(f'DISCOUNT FACTOR (MARR {rate_percent}%)')
    cash_flow_data['Total'].append('')
    for i, year in enumerate(years_labels):
        cash_flow_data[year].append(discount_factors[i])

    # Discounted Cash Flow
    cash_flow_data['Description'].append('DISCOUNTED CASH FLOW')
    cash_flow_data['Total'].append('')
    for i, year in enumerate(years_labels):
        cash_flow_data[year].append(discounted_cashflows[i])

    # Cumulative Cash Flow
    cash_flow_data['Description'].append('CUMULATIVE CASH FLOW')
    cash_flow_data['Total'].append('')
    for i, year in enumerate(years_labels):
        cash_flow_data[year].append(cumulative_cashflows[i])

    # Create DataFrame
    df_cashflow = pd.DataFrame(cash_flow_data)

    # Format the dataframe for display with improved formatting
    def format_cell(x):
        if isinstance(x, (int, float)) and x != '':
            if abs(x) >= 1e6:
                return f"Rp {x/1e6:.2f}M"
            elif abs(x) >= 1e3:
                return f"Rp {x/1e3:.0f}K"
            else:
                return f"Rp {x:,.0f}"
        return str(x) if x != '' else ''

    df_display = df_cashflow.copy()
    for col in df_display.columns:
        if col != 'Description':
            df_display[col] = df_display[col].apply(format_cell)

    # Display table with enhanced styling
    st.dataframe(
        df_display,
        use_container_width=True,
        height=800,
        hide_index=True
    )

    # Enhanced download buttons
    col1, col2 = st.columns(2)

    with col1:
        st.download_button(
            label="📥 Download Cash Flow Table (CSV)",
            data=df_cashflow.to_csv(index=False),
            file_name=f"cashflow_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            mime="text/csv"
        )

    with col2:
        excel_file = export_to_excel()
        if excel_file:
            st.download_button(
                label="📥 Download Excel (Formatted)",
                data=excel_file,
                file_name=f"feasibility_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# TAB 3: FINANCIAL METRICS (Enhanced)
with tab3:
    st.markdown('<div class="section-header">Financial Metrics & Analysis</div>', unsafe_allow_html=True)

    # Calculate metrics
    cashflows = calculate_net_cashflow()
    discount_factors = calculate_discount_factor()
    discounted_cashflows = calculate_discounted_cashflow(cashflows, discount_factors)
    cumulative_cashflows = calculate_cumulative_cashflow(discounted_cashflows)

    npv = calculate_npv(discounted_cashflows)
    irr = calculate_irr(cashflows)
    pbp = calculate_payback_period(cumulative_cashflows)

    # Display metrics in enhanced cards
    col1, col2, col3 = st.columns(3)

    with col1:
        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
        st.metric(
            label="Net Present Value (NPV)",
            value=f"Rp {npv:,.0f}",
            delta="Positive" if npv > 0 else "Negative",
            delta_color="normal" if npv > 0 else "inverse"
        )
        if npv > 0:
            st.success("✅ Project is FEASIBLE")
        else:
            st.error("❌ Project is NOT FEASIBLE")
        st.markdown('</div>', unsafe_allow_html=True)

    with col2:
        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
        st.metric(
            label="Internal Rate of Return (IRR)",
            value=f"{irr*100:.2f}%",
            delta=f"{(irr*100 - st.session_state.discount_rate):.2f}% vs MARR",
            delta_color="normal" if irr*100 > st.session_state.discount_rate else "inverse"
        )
        if irr * 100 > st.session_state.discount_rate:
            st.success("✅ IRR > MARR (Good)")
        else:
            st.warning("⚠️ IRR < MARR (Risky)")
        st.markdown('</div>', unsafe_allow_html=True)

    with col3:
        st.markdown('<div class="metric-card">', unsafe_allow_html=True)
        st.metric(
            label="Payback Period (PBP)",
            value=f"{pbp:.2f} years",
        )
        if pbp <= st.session_state.project_years:
            st.success(f"✅ Payback within project duration")
        else:
            st.warning(f"⚠️ Payback exceeds project duration")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")

    # Additional Details
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Investment Summary")
        st.write(f"**Initial Investment (CAPEX):** Rp {-cashflows[0]:,.0f}")
        st.write(f"**Annual Net Operating Cash Flow:** Rp {cashflows[1]:,.0f}")
        st.write(f"**Project Duration:** {st.session_state.project_years} years")
        st.write(f"**Discount Rate (MARR):** {st.session_state.discount_rate}%")

    with col2:
        st.subheader("Decision Criteria")
        criteria_data = {
            'Metric': ['NPV', 'IRR', 'Payback Period'],
            'Value': [
                f"Rp {npv:,.0f}",
                f"{irr*100:.2f}%",
                f"{pbp:.2f} years"
            ],
            'Criteria': [
                'NPV > 0',
                f'IRR > {st.session_state.discount_rate}%',
                f'PBP < {st.session_state.project_years} years'
            ],
            'Status': [
                '✅ Pass' if npv > 0 else '❌ Fail',
                '✅ Pass' if irr*100 > st.session_state.discount_rate else '❌ Fail',
                '✅ Pass' if pbp <= st.session_state.project_years else '❌ Fail'
            ]
        }
        st.dataframe(criteria_data, use_container_width=True, hide_index=True)

    st.markdown("---")

    # Interpretation
    st.subheader("📋 Interpretation & Recommendation")

    passing_criteria = sum([
        npv > 0,
        irr * 100 > st.session_state.discount_rate,
        pbp <= st.session_state.project_years
    ])

    if passing_criteria == 3:
        st.success("""
        ### ✅ STRONG RECOMMENDATION: PROCEED WITH PROJECT

        All three financial criteria are met:
        - **NPV is positive**: The project will create value
        - **IRR exceeds MARR**: Returns exceed the required rate
        - **Payback period is acceptable**: Investment will be recovered within project timeline

        This project demonstrates strong financial viability and should be considered for implementation.
        """)
    elif passing_criteria == 2:
        st.warning("""
        ### ⚠️ CONDITIONAL RECOMMENDATION: PROCEED WITH CAUTION

        Two out of three financial criteria are met. Consider:
        - Reviewing assumptions and inputs
        - Conducting sensitivity analysis
        - Evaluating non-financial factors
        - Exploring optimization opportunities
        """)
    else:
        st.error("""
        ### ❌ NOT RECOMMENDED: REJECT OR REDESIGN PROJECT

        The project fails to meet most financial criteria. Recommendations:
        - Reconsider the project scope and scale
        - Explore cost reduction opportunities
        - Investigate revenue enhancement strategies
        - Consider alternative investment options
        """)

# TAB 4: VISUALIZATIONS (Enhanced)
with tab4:
    st.markdown('<div class="section-header">Visualizations & Charts</div>', unsafe_allow_html=True)

    # Calculate values
    cashflows = calculate_net_cashflow()
    discount_factors = calculate_discount_factor()
    discounted_cashflows = calculate_discounted_cashflow(cashflows, discount_factors)
    cumulative_cashflows = calculate_cumulative_cashflow(discounted_cashflows)
    years_labels = ['Year 0'] + [f'Year {i+1}' for i in range(int(st.session_state.project_years))]

    # 1. Cash Flow Chart
    st.subheader("💰 Net Cash Flow by Year")
    fig_cashflow = go.Figure()
    fig_cashflow.add_trace(go.Bar(
        x=years_labels,
        y=cashflows,
        name='Net Cash Flow',
        marker_color=['red' if cf < 0 else 'green' for cf in cashflows],
        text=[f'Rp {cf:,.0f}' for cf in cashflows],
        textposition='outside'
    ))
    fig_cashflow.update_layout(
        title='Net Cash Flow Analysis',
        xaxis_title='Year',
        yaxis_title='Cash Flow (Rp)',
        hovermode='x unified',
        height=400
    )
    st.plotly_chart(fig_cashflow, use_container_width=True)

    # 2. Cumulative Cash Flow Chart
    st.subheader("📈 Cumulative Discounted Cash Flow")
    fig_cumulative = go.Figure()
    fig_cumulative.add_trace(go.Scatter(
        x=years_labels,
        y=cumulative_cashflows,
        mode='lines+markers',
        name='Cumulative Cash Flow',
        line=dict(color='blue', width=3),
        marker=dict(size=10),
        fill='tozeroy'
    ))
    fig_cumulative.add_hline(y=0, line_dash="dash", line_color="red", annotation_text="Break-even")
    fig_cumulative.update_layout(
        title='Cumulative Discounted Cash Flow (NPV Progression)',
        xaxis_title='Year',
        yaxis_title='Cumulative Cash Flow (Rp)',
        hovermode='x unified',
        height=400
    )
    st.plotly_chart(fig_cumulative, use_container_width=True)

    # 3. CAPEX Breakdown
    st.subheader("💼 CAPEX Breakdown")
    capex_names = [item['name'] for item in st.session_state.capex_items]
    capex_values = [calculate_total(item) for item in st.session_state.capex_items]

    fig_capex = go.Figure(data=[go.Pie(
        labels=capex_names,
        values=capex_values,
        hole=0.4,
        textinfo='label+percent',
        hovertemplate='<b>%{label}</b><br>Rp %{value:,.0f}<br>%{percent}<extra></extra>'
    )])
    fig_capex.update_layout(
        title='Capital Expenditure Distribution',
        height=400
    )
    st.plotly_chart(fig_capex, use_container_width=True)

    # 4. OPEX Analysis
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("💵 Revenue Sources")
        revenue_names = [item['name'] for item in st.session_state.opex_cash_in]
        revenue_values = [calculate_total(item) for item in st.session_state.opex_cash_in]

        fig_revenue = go.Figure(data=[go.Bar(
            y=revenue_names,
            x=revenue_values,
            orientation='h',
            marker_color='lightgreen',
            text=[f'Rp {v:,.0f}' for v in revenue_values],
            textposition='outside'
        )])
        fig_revenue.update_layout(
            title='Annual Revenue Breakdown',
            xaxis_title='Amount (Rp)',
            height=300
        )
        st.plotly_chart(fig_revenue, use_container_width=True)

    with col2:
        st.subheader("💸 Expense Categories")
        expense_names = [item['name'] for item in st.session_state.opex_cash_out]
        expense_values = [calculate_total(item) for item in st.session_state.opex_cash_out]

        fig_expense = go.Figure(data=[go.Bar(
            y=expense_names,
            x=expense_values,
            orientation='h',
            marker_color='lightcoral',
            text=[f'Rp {v:,.0f}' for v in expense_values],
            textposition='outside'
        )])
        fig_expense.update_layout(
            title='Annual Expense Breakdown',
            xaxis_title='Amount (Rp)',
            height=300
        )
        st.plotly_chart(fig_expense, use_container_width=True)

    # 5. Financial Metrics Comparison
    st.subheader("📊 Key Financial Metrics")

    npv = calculate_npv(discounted_cashflows)
    irr = calculate_irr(cashflows)
    pbp = calculate_payback_period(cumulative_cashflows)

    metrics_data = {
        'Metric': ['NPV (Million Rp)', 'IRR (%)', 'Payback Period (Years)'],
        'Value': [npv/1000000, irr*100, pbp],
        'Target': [0, st.session_state.discount_rate, st.session_state.project_years]
    }

    fig_metrics = go.Figure()
    fig_metrics.add_trace(go.Bar(
        name='Actual',
        x=metrics_data['Metric'],
        y=metrics_data['Value'],
        marker_color=['green' if v > t else 'red' for v, t in zip(metrics_data['Value'], metrics_data['Target'])]
    ))
    fig_metrics.add_trace(go.Scatter(
        name='Target/Threshold',
        x=metrics_data['Metric'],
        y=metrics_data['Target'],
        mode='markers',
        marker=dict(size=15, symbol='line-ew-open', color='orange', line=dict(width=3))
    ))
    fig_metrics.update_layout(
        title='Financial Metrics vs Targets',
        yaxis_title='Value',
        height=400,
        barmode='group'
    )
    st.plotly_chart(fig_metrics, use_container_width=True)

# TAB 5: SENSITIVITY ANALYSIS (Enhanced)
with tab5:
    st.markdown('<div class="section-header">Sensitivity Analysis (Tornado Diagram)</div>', unsafe_allow_html=True)

    st.info("""
    🎯 **Sensitivity Analysis** shows how changes in key variables affect your project's NPV.

    **What we test:** How NPV changes when each variable changes by the same percentage (±X%).

    **Why it matters:** Helps identify which variables deserve the most attention and risk management.
    """)

    # Sensitivity parameters
    variation = st.slider("Variation Percentage (%)", min_value=5, max_value=50, value=20, step=5)
    variation_decimal = variation / 100

    # Base case NPV
    base_cashflows = calculate_net_cashflow()
    base_discount_factors = calculate_discount_factor()
    base_discounted = calculate_discounted_cashflow(base_cashflows, base_discount_factors)
    base_npv = calculate_npv(base_discounted)

    # Variables to test
    sensitivity_results = []

    # 1. Sensitivity to Revenue
    st.session_state.opex_cash_in_backup = st.session_state.opex_cash_in.copy()

    # Increase revenue
    for item in st.session_state.opex_cash_in:
        item['price'] = item['price'] * (1 + variation_decimal)
    cashflows_high = calculate_net_cashflow()
    discounted_high = calculate_discounted_cashflow(cashflows_high, base_discount_factors)
    npv_high = calculate_npv(discounted_high)

    # Decrease revenue
    for item in st.session_state.opex_cash_in:
        item['price'] = item['price'] * (1 - variation_decimal)
    cashflows_low = calculate_net_cashflow()
    discounted_low = calculate_discounted_cashflow(cashflows_low, base_discount_factors)
    npv_low = calculate_npv(discounted_low)

    # Restore
    st.session_state.opex_cash_in = st.session_state.opex_cash_in_backup

    sensitivity_results.append({
        'Variable': 'Revenue',
        'NPV_Low': npv_low,
        'NPV_High': npv_high,
        'Range': npv_high - npv_low
    })

    # 2. Sensitivity to Operating Costs
    st.session_state.opex_cash_out_backup = st.session_state.opex_cash_out.copy()

    # Increase costs (decreases NPV)
    for item in st.session_state.opex_cash_out:
        item['price'] = item['price'] * (1 + variation_decimal)
    cashflows_high = calculate_net_cashflow()
    discounted_high = calculate_discounted_cashflow(cashflows_high, base_discount_factors)
    npv_cost_high = calculate_npv(discounted_high)

    # Decrease costs (increases NPV)
    for item in st.session_state.opex_cash_out:
        item['price'] = item['price'] * (1 - variation_decimal)
    cashflows_low = calculate_net_cashflow()
    discounted_low = calculate_discounted_cashflow(cashflows_low, base_discount_factors)
    npv_cost_low = calculate_npv(discounted_low)

    # Restore
    st.session_state.opex_cash_out = st.session_state.opex_cash_out_backup

    sensitivity_results.append({
        'Variable': 'Operating Costs',
        'NPV_Low': npv_cost_high,  # Note: swapped because higher cost = lower NPV
        'NPV_High': npv_cost_low,
        'Range': npv_cost_low - npv_cost_high
    })

    # 3. Sensitivity to Initial Investment (CAPEX)
    st.session_state.capex_items_backup = st.session_state.capex_items.copy()

    # Increase CAPEX
    for item in st.session_state.capex_items:
        item['price'] = item['price'] * (1 + variation_decimal)
    cashflows_high = calculate_net_cashflow()
    discounted_high = calculate_discounted_cashflow(cashflows_high, base_discount_factors)
    npv_capex_high = calculate_npv(discounted_high)

    # Decrease CAPEX
    for item in st.session_state.capex_items:
        item['price'] = item['price'] * (1 - variation_decimal)
    cashflows_low = calculate_net_cashflow()
    discounted_low = calculate_discounted_cashflow(cashflows_low, base_discount_factors)
    npv_capex_low = calculate_npv(discounted_low)

    # Restore
    st.session_state.capex_items = st.session_state.capex_items_backup

    sensitivity_results.append({
        'Variable': 'Initial Investment',
        'NPV_Low': npv_capex_high,  # Note: swapped
        'NPV_High': npv_capex_low,
        'Range': npv_capex_low - npv_capex_high
    })

    # 4. Sensitivity to Discount Rate
    original_rate = st.session_state.discount_rate

    # Increase discount rate
    st.session_state.discount_rate = original_rate * (1 + variation_decimal)
    df_high = calculate_discount_factor()
    discounted_high = calculate_discounted_cashflow(base_cashflows, df_high)
    npv_rate_high = calculate_npv(discounted_high)

    # Decrease discount rate
    st.session_state.discount_rate = original_rate * (1 - variation_decimal)
    df_low = calculate_discount_factor()
    discounted_low = calculate_discounted_cashflow(base_cashflows, df_low)
    npv_rate_low = calculate_npv(discounted_low)

    # Restore
    st.session_state.discount_rate = original_rate

    sensitivity_results.append({
        'Variable': 'Discount Rate',
        'NPV_Low': npv_rate_high,  # Note: swapped
        'NPV_High': npv_rate_low,
        'Range': npv_rate_low - npv_rate_high
    })

    # Sort by range (most sensitive first)
    sensitivity_results.sort(key=lambda x: x['Range'], reverse=True)

    # Create Tornado Diagram
    fig_tornado = go.Figure()

    variables = [r['Variable'] for r in sensitivity_results]
    low_values = [r['NPV_Low'] - base_npv for r in sensitivity_results]
    high_values = [r['NPV_High'] - base_npv for r in sensitivity_results]

    # Add bars for negative impact (adverse scenario)
    fig_tornado.add_trace(go.Bar(
        name=f'Worst Case (-{variation}%)',
        y=variables,
        x=low_values,
        orientation='h',
        marker=dict(color='#FF6B6B'),
        hovertemplate='<b>%{y}</b><br>NPV Change: Rp %{x:,.0f}<br>Scenario: Worst Case<extra></extra>'
    ))

    # Add bars for positive impact (best scenario)
    fig_tornado.add_trace(go.Bar(
        name=f'Best Case (+{variation}%)',
        y=variables,
        x=high_values,
        orientation='h',
        marker=dict(color='#4ECDC4'),
        hovertemplate='<b>%{y}</b><br>NPV Change: Rp %{x:,.0f}<br>Scenario: Best Case<extra></extra>'
    ))

    # Add baseline
    fig_tornado.add_vline(x=0, line_dash="dash", line_color="black", line_width=2,
                          annotation_text="Base Case NPV")

    # Improve layout
    fig_tornado.update_layout(
        title=f'<b>Tornado Diagram - Sensitivity Analysis</b><br><sub>Impact of ±{variation}% changes on NPV</sub>',
        xaxis_title=f'NPV Change from Base Case (Rp {base_npv:,.0f})',
        yaxis_title='Variables (Ranked by Impact)',
        barmode='overlay',
        height=500,
        showlegend=True,
        legend=dict(x=0.7, y=0.1),
        template='plotly_white'
    )

    st.plotly_chart(fig_tornado, use_container_width=True)

    # Display sensitivity table
    st.subheader("📋 Sensitivity Analysis Summary")

    sensitivity_table = {
        'Variable': [r['Variable'] for r in sensitivity_results],
        f'NPV at -{variation}%': [f"Rp {r['NPV_Low']:,.0f}" for r in sensitivity_results],
        'Base NPV': [f"Rp {base_npv:,.0f}" for _ in sensitivity_results],
        f'NPV at +{variation}%': [f"Rp {r['NPV_High']:,.0f}" for r in sensitivity_results],
        'NPV Range': [f"Rp {r['Range']:,.0f}" for r in sensitivity_results],
        'Sensitivity Rank': list(range(1, len(sensitivity_results) + 1))
    }

    st.dataframe(sensitivity_table, use_container_width=True, hide_index=True)

    # Enhanced interpretation
    most_sensitive = sensitivity_results[0]
    least_sensitive = sensitivity_results[-1]

    col1, col2 = st.columns(2)

    with col1:
        st.markdown(f"""
        ### 🎯 Key Insights
        - **Most Critical Variable:** `{most_sensitive['Variable']}`
        - **Potential NPV Impact:** Rp {most_sensitive['Range']:,.0f}
        - **Risk Level:** {'🔴 High' if most_sensitive['Range'] > abs(base_npv * 0.5) else '🟡 Medium' if most_sensitive['Range'] > abs(base_npv * 0.2) else '🟢 Low'}
        """)

    with col2:
        st.markdown(f"""
        ### 📊 Recommendations
        - **Monitor Closely:** {most_sensitive['Variable']}
        - **Priority:** High attention needed
        - **Action:** Implement risk mitigation strategies
        """)

    # Detailed explanation
    with st.expander("📖 How to Interpret This Analysis"):
        st.markdown(f"""
        ### Understanding the Tornado Diagram

        **What it shows:**
        - How much NPV changes when each variable changes by ±{variation}%
        - Variables are ranked by impact (most sensitive at top)
        - The wider the bar, the more sensitive the NPV is to that variable

        **Base Case NPV:** Rp {base_npv:,.0f}

        **Scenarios Explained:**
        - **Worst Case (-{variation}%):** When variable decreases (for revenue) or increases (for costs)
        - **Best Case (+{variation}%):** When variable increases (for revenue) or decreases (for costs)

        **Management Implications:**
        1. **{most_sensitive['Variable']}** deserves the most attention and monitoring
        2. Small changes in {most_sensitive['Variable']} can significantly impact project viability
        3. Consider conducting more detailed analysis or implementing risk mitigation for top 3 variables

        **Next Steps:**
        - Gather more accurate data for sensitive variables
        - Develop contingency plans for adverse scenarios
        - Consider sensitivity analysis results in your go/no-go decision
        """)

# Footer
st.markdown("---")
st.markdown("""
    <div style='text-align: center; color: gray; padding: 1rem;'>
        <p>Feasibility Analysis Tool v2.0 | Built with Streamlit</p>
        <p>For investment analysis and project evaluation | Enhanced with improved sensitivity analysis</p>
    </div>
    """, unsafe_allow_html=True)